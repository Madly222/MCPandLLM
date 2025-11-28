import os
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
STORAGE_DIR = Path(os.getenv("FILES_DIR", BASE_DIR / "storage"))
DOWNLOADS_DIR = Path(os.getenv("DOWNLOADS_DIR", BASE_DIR / "downloads"))
SERVER_URL = os.getenv("SERVER_URL", "http://localhost:8000")

DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

MAX_FILE_AGE_MINUTES = 5


def _cleanup_old_downloads():
    if not DOWNLOADS_DIR.exists():
        return

    now = datetime.now()
    for filepath in DOWNLOADS_DIR.iterdir():
        if filepath.is_file() and filepath.suffix.lower() in ['.xlsx', '.xls']:
            file_age = now - datetime.fromtimestamp(filepath.stat().st_mtime)
            if file_age > timedelta(minutes=MAX_FILE_AGE_MINUTES):
                try:
                    filepath.unlink()
                    logger.info(f"Удалён старый файл: {filepath.name}")
                except Exception as e:
                    logger.error(f"Ошибка удаления {filepath.name}: {e}")


def _find_source_file(filename: str) -> Optional[Path]:
    for directory in [STORAGE_DIR, DOWNLOADS_DIR]:
        path = directory / filename
        if path.exists():
            logger.info(f"Найден файл: {path}")
            return path

    filename_clean = filename.lower().replace(' ', '').replace('(', '').replace(')', '')

    for directory in [STORAGE_DIR, DOWNLOADS_DIR]:
        if not directory.exists():
            continue
        for filepath in directory.iterdir():
            if filepath.suffix.lower() in ['.xlsx', '.xls']:
                name_clean = filepath.name.lower().replace(' ', '').replace('(', '').replace(')', '')
                if filename_clean == name_clean:
                    logger.info(f"Найден файл по fuzzy: {filepath}")
                    return filepath

                if filename_clean.replace('.xlsx', '').replace('.xls', '') in name_clean:
                    logger.info(f"Найден файл по частичному совпадению: {filepath}")
                    return filepath

    logger.error(f"Файл не найден: {filename}")
    logger.error(f"STORAGE_DIR: {STORAGE_DIR}, существует: {STORAGE_DIR.exists()}")
    if STORAGE_DIR.exists():
        logger.error(f"Файлы в storage: {[f.name for f in STORAGE_DIR.iterdir()]}")

    return None


def _generate_output_filename(original: str) -> str:
    stem = Path(original).stem
    suffix = Path(original).suffix
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}_edited_{timestamp}{suffix}"


def _copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def edit_excel(
        filename: str,
        operations: List[Dict[str, Any]],
        sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    Редактирует Excel файл и возвращает ссылку на скачивание.

    Операции:
    - {"action": "add_row", "data": ["val1", "val2", ...], "after_row": 5}
    - {"action": "edit_cell", "row": 5, "col": 2, "value": "новое значение"}
    - {"action": "edit_cell", "row": 5, "col": "B", "value": "новое значение"}
    - {"action": "delete_row", "row": 5}
    - {"action": "add_column", "header": "Новая колонка", "after_col": 3}
    - {"action": "delete_column", "col": 3}

    Возвращает:
    {
        "success": True,
        "download_url": "http://localhost:8000/download/file_edited_20250115.xlsx",
        "filename": "file_edited_20250115.xlsx",
        "operations_applied": 3
    }
    """
    _cleanup_old_downloads()

    source_path = _find_source_file(filename)
    if not source_path:
        return {
            "success": False,
            "error": f"Файл {filename} не найден"
        }

    try:
        wb = load_workbook(source_path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Лист '{sheet_name}' не найден. Доступные: {wb.sheetnames}"
                }
            ws = wb[sheet_name]
        else:
            ws = wb.active

        ops_applied = 0

        for op in operations:
            action = op.get("action", "").lower()

            if action == "add_row":
                data = op.get("data", [])
                after_row = op.get("after_row", ws.max_row)

                ws.insert_rows(after_row + 1)
                new_row = after_row + 1

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=new_row, column=col_idx, value=value)
                    if after_row > 0:
                        source_cell = ws.cell(row=after_row, column=col_idx)
                        _copy_cell_style(source_cell, cell)

                ops_applied += 1
                logger.info(f"Добавлена строка {new_row}: {data}")

            elif action == "edit_cell":
                row = op.get("row")
                col = op.get("col")
                value = op.get("value")

                if isinstance(col, str):
                    from openpyxl.utils import column_index_from_string
                    col = column_index_from_string(col)

                if row and col:
                    ws.cell(row=row, column=col, value=value)
                    ops_applied += 1
                    logger.info(f"Изменена ячейка ({row}, {col}): {value}")

            elif action == "delete_row":
                row = op.get("row")
                if row:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).value = None
                    ops_applied += 1
                    logger.info(f"Очищена строка {row}")

            elif action == "add_column":
                header = op.get("header", "")
                after_col = op.get("after_col", ws.max_column)

                ws.insert_cols(after_col + 1)
                new_col = after_col + 1
                ws.cell(row=1, column=new_col, value=header)
                ops_applied += 1
                logger.info(f"Добавлена колонка {new_col}: {header}")

            elif action == "delete_column":
                col = op.get("col")
                if isinstance(col, str):
                    from openpyxl.utils import column_index_from_string
                    col = column_index_from_string(col)
                if col:
                    ws.delete_cols(col)
                    ops_applied += 1
                    logger.info(f"Удалена колонка {col}")

            else:
                logger.warning(f"Неизвестная операция: {action}")

        output_filename = _generate_output_filename(filename)
        output_path = DOWNLOADS_DIR / output_filename

        wb.calculation = CalcProperties(fullCalcOnLoad=True)

        wb.save(output_path)
        wb.close()

        download_url = f"{SERVER_URL}/download/{output_filename}"

        logger.info(f"Файл сохранён: {output_path}")

        return {
            "success": True,
            "download_url": download_url,
            "filename": output_filename,
            "operations_applied": ops_applied,
            "message": f"Файл отредактирован. Скачать: {download_url}"
        }

    except Exception as e:
        logger.error(f"Ошибка редактирования {filename}: {e}")
        return {
            "success": False,
            "error": str(e)
        }


def add_row_to_excel(
        filename: str,
        row_data: List[Any],
        after_row: Optional[int] = None,
        sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """Упрощённая функция добавления строки"""
    op = {"action": "add_row", "data": row_data}
    if after_row:
        op["after_row"] = after_row
    return edit_excel(filename, [op], sheet_name)


def edit_cell_in_excel(
        filename: str,
        row: int,
        col: Any,
        value: Any,
        sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """Упрощённая функция изменения ячейки"""
    return edit_excel(filename, [{"action": "edit_cell", "row": row, "col": col, "value": value}], sheet_name)


def delete_row_from_excel(
        filename: str,
        row: int,
        sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """Упрощённая функция удаления строки"""
    return edit_excel(filename, [{"action": "delete_row", "row": row}], sheet_name)


def get_excel_preview(filename: str, rows: int = 10, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Получить превью таблицы для LLM"""
    source_path = _find_source_file(filename)
    if not source_path:
        return {"success": False, "error": f"Файл {filename} не найден"}

    try:
        wb = load_workbook(source_path, data_only=True)

        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        else:
            ws = wb.active

        data = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=rows + 1, values_only=True), start=1):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            data.append({"row": row_idx, "values": row_data})

        wb.close()

        return {
            "success": True,
            "filename": filename,
            "sheet": ws.title,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "preview": data
        }

    except Exception as e:
        return {"success": False, "error": str(e)}