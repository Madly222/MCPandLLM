import os
from typing import Optional, Dict, List, Tuple, Union
from pathlib import Path
from openpyxl import load_workbook, Workbook

BASE_DIR = Path(__file__).resolve().parent.parent
STORAGE_DIR = Path(os.getenv("FILES_DIR", BASE_DIR / "storage"))


def _resolve_path(filename: Union[str, Path], role: Optional[str] = None) -> Path:
    if isinstance(filename, Path):
        return filename

    if role:
        path = STORAGE_DIR / role / filename
        if path.exists():
            return path

    path = STORAGE_DIR / filename
    if path.exists():
        return path

    for role_dir in STORAGE_DIR.iterdir():
        if role_dir.is_dir():
            path = role_dir / filename
            if path.exists():
                return path

    return STORAGE_DIR / filename


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _detect_headers(rows: List[List[str]]) -> Tuple[Optional[List[str]], int]:
    for idx, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c]
        if len(non_empty) >= 2:
            has_text = sum(1 for c in non_empty if not c.replace(".", "").replace(",", "").replace("-", "").isdigit())
            if has_text >= len(non_empty) * 0.5:
                return row, idx
    return None, 0


def _to_markdown_table_with_row_numbers(
        headers: List[str],
        data_rows: List[List[str]],
        start_row: int = 2,
        max_rows: int = 500
) -> str:
    if not headers:
        return ""

    col_count = len(headers)
    clean_headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]

    lines = []
    lines.append("| ROW | " + " | ".join(clean_headers) + " |")
    lines.append("| --- | " + " | ".join(["---"] * col_count) + " |")

    for i, row in enumerate(data_rows[:max_rows]):
        excel_row = start_row + i
        cells = []
        for j in range(col_count):
            val = row[j] if j < len(row) else ""
            cells.append(val.replace("|", "/").replace("\n", " "))
        lines.append(f"| {excel_row} | " + " | ".join(cells) + " |")

    if len(data_rows) > max_rows:
        lines.append(f"\n... и ещё {len(data_rows) - max_rows} строк")

    return "\n".join(lines)


def _to_markdown_table(headers: List[str], data_rows: List[List[str]], max_rows: int = 500) -> str:
    if not headers:
        return ""

    col_count = len(headers)
    clean_headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]

    lines = []
    lines.append("| " + " | ".join(clean_headers) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")

    for row in data_rows[:max_rows]:
        cells = []
        for i in range(col_count):
            val = row[i] if i < len(row) else ""
            cells.append(val.replace("|", "/").replace("\n", " "))
        lines.append("| " + " | ".join(cells) + " |")

    if len(data_rows) > max_rows:
        lines.append(f"\n... и ещё {len(data_rows) - max_rows} строк")

    return "\n".join(lines)


def _extract_structure(headers: List[str], data_rows: List[List[str]]) -> Dict:
    structure = {
        "columns": [h for h in headers if h],
        "column_count": len([h for h in headers if h]),
        "row_count": len(data_rows),
        "sample_values": {}
    }

    for i, header in enumerate(headers):
        if not header:
            continue
        values = []
        for row in data_rows[:10]:
            if i < len(row) and row[i]:
                values.append(row[i])
        if values:
            structure["sample_values"][header] = values[:3]

    return structure


def _generate_summary(filename: str, sheets_info: List[Dict]) -> str:
    parts = [f"Файл: {filename}"]

    for sheet in sheets_info:
        sheet_desc = f"Лист '{sheet['name']}': {sheet['row_count']} строк"
        if sheet.get("columns"):
            cols = ", ".join(sheet["columns"][:10])
            if len(sheet["columns"]) > 10:
                cols += f" ... (+{len(sheet['columns']) - 10})"
            sheet_desc += f". Колонки: {cols}"
        parts.append(sheet_desc)

    return ". ".join(parts)


def read_excel(filename: Union[str, Path], with_row_numbers: bool = False, role: Optional[str] = None) -> str:
    path = _resolve_path(filename, role)

    if not path.exists():
        return f"Файл '{filename}' не найден."

    try:
        wb = load_workbook(path, data_only=True)
        parts = []
        display_name = path.name if isinstance(filename, Path) else filename
        parts.append(f"# {display_name}\n")

        for sheet in wb.worksheets:
            parts.append(f"\n## Лист: {sheet.title}\n")

            all_rows = []
            row_numbers = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [_format_cell(cell) for cell in row]
                if any(cell for cell in cells):
                    all_rows.append(cells)
                    row_numbers.append(row_idx)

            if not all_rows:
                parts.append("(пустой лист)\n")
                continue

            headers, header_idx = _detect_headers(all_rows)

            if header_idx > 0:
                for i, row in enumerate(all_rows[:header_idx]):
                    info = " | ".join(c for c in row if c)
                    if info:
                        if with_row_numbers:
                            parts.append(f"**[ROW {row_numbers[i]}] {info}**\n")
                        else:
                            parts.append(f"**{info}**\n")

            if headers:
                data_rows = all_rows[header_idx + 1:]
                start_row = row_numbers[header_idx + 1] if header_idx + 1 < len(row_numbers) else header_idx + 2

                if with_row_numbers:
                    md_table = _to_markdown_table_with_row_numbers(headers, data_rows, start_row=start_row)
                else:
                    md_table = _to_markdown_table(headers, data_rows)
                parts.append(md_table)
            else:
                for i, row in enumerate(all_rows):
                    if with_row_numbers:
                        parts.append(f"[ROW {row_numbers[i]}] " + " | ".join(c for c in row if c) + "\n")
                    else:
                        parts.append(" | ".join(c for c in row if c) + "\n")

        return "\n".join(parts)

    except Exception as e:
        return f"Ошибка чтения Excel: {e}"


def read_excel_for_edit(filename: Union[str, Path], role: Optional[str] = None) -> str:
    return read_excel(filename, with_row_numbers=True, role=role)


def read_excel_structured(filename: Union[str, Path], role: Optional[str] = None) -> Dict:
    path = _resolve_path(filename, role)

    if not path.exists():
        return {"error": f"Файл '{filename}' не найден."}

    try:
        wb = load_workbook(path, data_only=True)
        display_name = path.name if isinstance(filename, Path) else filename

        result = {
            "filename": display_name,
            "sheets": [],
            "content": "",
            "summary": "",
            "structure": "",
            "total_rows": 0,
            "all_columns": []
        }

        content_parts = [f"# {display_name}\n"]
        sheets_info = []

        for sheet in wb.worksheets:
            sheet_data = {
                "name": sheet.title,
                "headers": [],
                "row_count": 0,
                "columns": [],
                "sample_values": {}
            }

            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [_format_cell(cell) for cell in row]
                if any(cell for cell in cells):
                    all_rows.append(cells)

            if not all_rows:
                sheets_info.append({"name": sheet.title, "row_count": 0, "columns": []})
                continue

            headers, header_idx = _detect_headers(all_rows)
            data_rows = all_rows[header_idx + 1:] if headers else all_rows

            content_parts.append(f"\n## Лист: {sheet.title}\n")

            if header_idx > 0:
                for row in all_rows[:header_idx]:
                    info = " | ".join(c for c in row if c)
                    if info:
                        content_parts.append(f"**{info}**\n")

            if headers:
                md_table = _to_markdown_table(headers, data_rows)
                content_parts.append(md_table)

                structure = _extract_structure(headers, data_rows)
                sheet_data["headers"] = headers
                sheet_data["row_count"] = len(data_rows)
                sheet_data["columns"] = structure["columns"]
                sheet_data["sample_values"] = structure["sample_values"]
                result["all_columns"].extend(structure["columns"])
            else:
                for row in all_rows:
                    content_parts.append(" | ".join(c for c in row if c) + "\n")
                sheet_data["row_count"] = len(all_rows)

            result["total_rows"] += sheet_data["row_count"]
            result["sheets"].append(sheet_data)
            sheets_info.append({
                "name": sheet.title,
                "row_count": sheet_data["row_count"],
                "columns": sheet_data.get("columns", [])
            })

        result["content"] = "\n".join(content_parts)
        result["summary"] = _generate_summary(display_name, sheets_info)
        result["structure"] = str({
            "sheets": [s["name"] for s in sheets_info],
            "total_rows": result["total_rows"],
            "columns": list(set(result["all_columns"]))
        })

        return result

    except Exception as e:
        return {"error": f"Ошибка чтения Excel: {e}"}


def write_excel(filename: str, sheet_name: str, data: list, role: Optional[str] = None) -> str:
    path = _resolve_path(filename, role)

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    for row in data:
        ws.append(row)

    wb.save(path)
    return f"Файл '{filename}' обновлён."