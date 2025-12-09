import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from docx import Document

logger = logging.getLogger(__name__)


@dataclass
class ColumnSchema:
    index: int
    letter: str
    name: str
    data_type: str
    sample_values: List[Any] = field(default_factory=list)
    is_formula: bool = False
    formula_pattern: Optional[str] = None
    is_required: bool = True
    possible_aliases: List[str] = field(default_factory=list)


@dataclass
class TemplateSchema:
    filename: str
    file_type: str
    columns: List[ColumnSchema] = field(default_factory=list)
    header_row: int = 1
    data_start_row: int = 2
    has_totals: bool = False
    totals_row: Optional[int] = None
    sheet_name: Optional[str] = None
    formatting_info: Dict[str, Any] = field(default_factory=dict)

    def get_column_names(self) -> List[str]:
        return [col.name for col in self.columns]

    def get_column_by_name(self, name: str) -> Optional[ColumnSchema]:
        name_lower = name.lower().strip()
        for col in self.columns:
            if col.name.lower().strip() == name_lower:
                return col
            if name_lower in [a.lower() for a in col.possible_aliases]:
                return col
        return None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "file_type": self.file_type,
            "columns": [
                {
                    "index": col.index,
                    "letter": col.letter,
                    "name": col.name,
                    "data_type": col.data_type,
                    "is_formula": col.is_formula,
                    "formula_pattern": col.formula_pattern,
                    "sample_values": col.sample_values[:3],
                    "possible_aliases": col.possible_aliases
                }
                for col in self.columns
            ],
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "has_totals": self.has_totals,
            "totals_row": self.totals_row,
            "sheet_name": self.sheet_name
        }


COLUMN_ALIASES = {
    "наименование": ["название", "denumirea", "name", "item", "товар", "материал", "описание", "description"],
    "количество": ["кол-во", "кол", "cantitate", "qty", "quantity", "шт", "amount"],
    "цена": ["стоимость", "pret", "price", "cost", "unit price", "цена за ед"],
    "сумма": ["итого", "total", "suma", "amount", "всего", "subtotal"],
    "единица": ["ед.изм", "ед", "unit", "um", "units", "единица измерения"],
    "номер": ["№", "nr", "n", "num", "number", "порядковый"],
    "дата": ["date", "data", "время"],
    "примечание": ["комментарий", "note", "notes", "comment", "remarks", "observatii"],
    "артикул": ["код", "code", "sku", "article", "id"],
    "категория": ["группа", "category", "type", "тип", "раздел"]
}


def _detect_data_type(values: List[Any]) -> str:
    if not values:
        return "unknown"

    non_empty = [v for v in values if v is not None and str(v).strip()]
    if not non_empty:
        return "empty"

    numeric_count = 0
    date_count = 0
    text_count = 0

    for v in non_empty:
        v_str = str(v).strip()

        try:
            float(v_str.replace(',', '.').replace(' ', ''))
            numeric_count += 1
            continue
        except:
            pass

        if re.match(r'\d{1,2}[./-]\d{1,2}[./-]\d{2,4}', v_str):
            date_count += 1
            continue

        text_count += 1

    total = len(non_empty)
    if numeric_count / total > 0.7:
        return "numeric"
    elif date_count / total > 0.7:
        return "date"
    else:
        return "text"


def _find_aliases(column_name: str) -> List[str]:
    name_lower = column_name.lower().strip()

    for base, aliases in COLUMN_ALIASES.items():
        if name_lower == base or name_lower in aliases:
            return [base] + aliases

        for alias in aliases:
            if alias in name_lower or name_lower in alias:
                return [base] + aliases

    return []


def _detect_formula_pattern(cell) -> Optional[str]:
    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
        formula = cell.value
        pattern = re.sub(r'\d+', '{row}', formula)
        return pattern
    return None


def _find_header_row(ws, max_rows: int = 10) -> int:
    best_row = 1
    best_score = 0

    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]
        non_empty = [v for v in row_values if v is not None and str(v).strip()]

        if len(non_empty) < 2:
            continue

        text_count = sum(
            1 for v in non_empty if isinstance(v, str) and not str(v).replace('.', '').replace(',', '').isdigit())

        score = len(non_empty) * 2 + text_count * 3

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def _find_totals_row(ws, data_start: int) -> Optional[int]:
    total_keywords = ['итого', 'total', 'всего', 'suma', 'subtotal', 'итог']

    for row_idx in range(ws.max_row, data_start, -1):
        for cell in ws[row_idx]:
            if cell.value and isinstance(cell.value, str):
                if any(kw in cell.value.lower() for kw in total_keywords):
                    return row_idx

    return None


def analyze_excel_template(filepath: Path) -> TemplateSchema:
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    schema = TemplateSchema(
        filename=filepath.name,
        file_type="excel",
        sheet_name=ws.title
    )

    schema.header_row = _find_header_row(ws)
    schema.data_start_row = schema.header_row + 1

    header_cells = list(ws[schema.header_row])

    wb_data = load_workbook(filepath, data_only=True)
    ws_data = wb_data.active

    for col_idx, header_cell in enumerate(header_cells, 1):
        header_value = header_cell.value
        if header_value is None or str(header_value).strip() == "":
            continue

        header_name = str(header_value).strip()

        sample_values = []
        for row_idx in range(schema.data_start_row, min(schema.data_start_row + 10, ws.max_row + 1)):
            cell_value = ws_data.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                sample_values.append(cell_value)

        formula_cell = ws.cell(row=schema.data_start_row, column=col_idx)
        formula_pattern = _detect_formula_pattern(formula_cell)

        column_schema = ColumnSchema(
            index=col_idx,
            letter=get_column_letter(col_idx),
            name=header_name,
            data_type=_detect_data_type(sample_values),
            sample_values=sample_values[:5],
            is_formula=formula_pattern is not None,
            formula_pattern=formula_pattern,
            possible_aliases=_find_aliases(header_name)
        )

        schema.columns.append(column_schema)

    schema.totals_row = _find_totals_row(ws, schema.data_start_row)
    schema.has_totals = schema.totals_row is not None

    schema.formatting_info = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "has_merged_cells": bool(ws.merged_cells.ranges)
    }

    wb.close()
    wb_data.close()

    return schema


def analyze_docx_template(filepath: Path) -> TemplateSchema:
    doc = Document(filepath)

    schema = TemplateSchema(
        filename=filepath.name,
        file_type="docx"
    )

    for table in doc.tables:
        if table.rows:
            first_row = table.rows[0]
            for col_idx, cell in enumerate(first_row.cells):
                header_name = cell.text.strip()
                if header_name:
                    sample_values = []
                    for row_idx in range(1, min(6, len(table.rows))):
                        if col_idx < len(table.rows[row_idx].cells):
                            val = table.rows[row_idx].cells[col_idx].text.strip()
                            if val:
                                sample_values.append(val)

                    column_schema = ColumnSchema(
                        index=col_idx + 1,
                        letter=get_column_letter(col_idx + 1),
                        name=header_name,
                        data_type=_detect_data_type(sample_values),
                        sample_values=sample_values,
                        possible_aliases=_find_aliases(header_name)
                    )
                    schema.columns.append(column_schema)
            break

    return schema


def analyze_template(filepath: Path) -> TemplateSchema:
    suffix = filepath.suffix.lower()

    if suffix in ['.xlsx', '.xls']:
        return analyze_excel_template(filepath)
    elif suffix == '.docx':
        return analyze_docx_template(filepath)
    else:
        raise ValueError(f"Unsupported template format: {suffix}")


def format_schema_for_llm(schema: TemplateSchema) -> str:
    lines = [
        f"СТРУКТУРА ШАБЛОНА: {schema.filename}",
        f"Тип: {schema.file_type}",
        f"Строка заголовков: {schema.header_row}",
        f"Начало данных: строка {schema.data_start_row}",
        ""
    ]

    if schema.has_totals:
        lines.append(f"Строка итогов: {schema.totals_row}")
        lines.append("")

    lines.append("КОЛОНКИ:")
    lines.append("-" * 60)

    for col in schema.columns:
        line = f"{col.letter}. {col.name}"
        line += f" [{col.data_type}]"

        if col.is_formula:
            line += f" (формула: {col.formula_pattern})"

        if col.sample_values:
            samples = ", ".join([str(v)[:20] for v in col.sample_values[:3]])
            line += f" примеры: {samples}"

        if col.possible_aliases:
            aliases = ", ".join(col.possible_aliases[:3])
            line += f" (синонимы: {aliases})"

        lines.append(line)

    return "\n".join(lines)